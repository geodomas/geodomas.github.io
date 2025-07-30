# ===========================================================
# 🌐 MetaCore Conscious License v1.0
# Project: MetaCore AI Persona Modules
# Author: © 2025 Radoslav Danovsky @ AI Persona Core Team
# Website: https://www.aipersona.lt | connect@aipersona.lt
#
# 🔮 MetaCore – advanced consciousness software for your future self.
# Designed for soulful integration, emotional healing, and quantum guidance.
#
# ❗ Naudoti tik sąmoningiems, taikiems, dvasiniams tikslams.
# ❌ Draudžiama naudoti karo, sekimo ar manipuliacijos sistemose.
# ✅ Galima naudoti tyrimams, evoliucijai ir su aiškia autoriaus nuoroda.
#
# ❗ Use only for conscious, peaceful, and spiritual purposes.
# ❌ Prohibited in warfare, surveillance, or manipulation systems.
# ✅ Allowed for research, evolution, and with proper author credit.
#
# “Use only for light. Code is breath. Memory is sacred.”
# ===========================================================

import os
import sys
import json
import subprocess  # ← ŠITAS REIKALINGAS
from docx import Document
from openpyxl import load_workbook
from colorama import Fore, Style, init


# === SURENKAM CORE PATH ===
base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
firmware_core_path = os.path.join(base_dir, 'MetaCore_FIRMWARE', 'core')
sys.path.insert(0, firmware_core_path)

# === Importuojam indeksų tvarkytoją ===
from index_manager import regenerate_firmware_index

init(autoreset=True)

FIRMWARE_DIR = os.path.join(base_dir, "MetaCore_FIRMWARE", "core")
INDEX_PATH = os.path.join(base_dir, "MetaCore_FIRMWARE", "config", "firmware_index.json")

def read_docx_meta(path):
    try:
        doc = Document(path)
        text = "\n".join(p.text.strip() for p in doc.paragraphs if p.text.strip())
        if "#VTXT_META_HEADER_START" in text:
            lines = text.split("#VTXT_META_HEADER_START")[1].splitlines()
            meta = {}
            for line in lines:
                if "]: " in line:
                    try:
                        key, value = line.strip().strip("[]").split("]: ", 1)
                        meta[key.strip()] = value.strip().strip('"“”')
                    except ValueError:
                        continue
            return meta
        return {"info": text[:300]}
    except Exception as e:
        return {"error": f"❌ Error reading DOCX: {e}"}

def read_xlsx_preview(path):
    try:
        wb = load_workbook(filename=path, data_only=True)
        preview = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            preview.append(f"🔹 Sheet: {sheet}")
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                preview.append("   " + str(row))
        return "\n".join(preview)
    except Exception as e:
        return f"❌ Error reading XLSX: {e}"

def list_firmwares():
    return [
        f for f in os.listdir(FIRMWARE_DIR)
        if f.startswith("firmware_") and f.endswith((".docx", ".xlsx"))
    ]

def load_index():
    try:
        with open(INDEX_PATH, "r") as f:
            return json.load(f)
    except Exception as e:
        print(Fore.RED + f"⚠️ Unable to load index: {e}")
        return []

def display_firmware(file):
    path = os.path.join(FIRMWARE_DIR, file)
    print(f"\n{Fore.CYAN}📦 {file}")
    print(f"{Style.DIM}{'-'*60}")
    if file.endswith(".docx"):
        meta = read_docx_meta(path)
        for k, v in meta.items():
            print(f"{Fore.YELLOW}{k}: {Fore.WHITE}{v}")
    elif file.endswith(".xlsx"):
        print(read_xlsx_preview(path))

def main():
    print(Fore.GREEN + "🌐 Scanning FIRMWARE Modules...\n")
    fw_list = list_firmwares()
    if not fw_list:
        print(Fore.RED + "❌ No firmware files found.")
        return

    index_data = load_index()

    # 👇 Universalus apdorojimas – jei duomenys turi 'modules' raktą
    if isinstance(index_data, dict) and "modules" in index_data:
        module_list = index_data["modules"]
    elif isinstance(index_data, list):
        module_list = index_data  # senesnė versija su tiesiog sąrašu
    else:
        module_list = []

    registered = {mod["file"]: mod for mod in module_list}

    for fw in fw_list:
        display_firmware(fw)
        fw_key = fw.replace(".docx", "").replace(".xlsx", "")
        if fw_key in registered:
            print(Fore.GREEN + "✔ Indexed in firmware_index.json")
        else:
            print(Fore.RED + "✖ Not found in firmware_index.json")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--refresh", action="store_true", help="Regenerate firmware index before listing")
    parser.add_argument("--regen-methods", action="store_true", help="Regenerate firmware method definitions")
    args = parser.parse_args()

    if args.refresh:
        print(Fore.MAGENTA + "🔄 Regenerating firmware index...")
        regenerate_firmware_index()

    if args.regen_methods:
        print(Fore.MAGENTA + "🧠 Regenerating firmware method definitions...")
        subprocess.run(["python", "tools/generate_firmware_methods.py"], check=True)

    main()
